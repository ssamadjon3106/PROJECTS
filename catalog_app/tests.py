from tempfile import NamedTemporaryFile

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from library.models import Book


def _create_tmp_xlsx(rows):
    """
    rows: list of (row_number_value, description_cell_value)
    Returns path to a temporary .xlsx file.
    """
    wb = Workbook()
    ws = wb.active

    for i, (row_number, description) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=row_number)
        ws.cell(row=i, column=2, value=description)

    tmp = NamedTemporaryFile(suffix='.xlsx', delete=False)
    try:
        wb.save(tmp.name)
        return tmp.name
    finally:
        tmp.close()


class ImportBooksXlsxTests(TestCase):
    def test_import_parses_float_row_index_and_isbn_with_label(self):
        tmp_path = _create_tmp_xlsx(
            [
                (
                    1.0,  # Excel often stores numeric cells as floats
                    'Книги Учебники, Clean Architecture, ISBN № 978-1-23456-789-7, Robert Martin',
                ),
            ]
        )

        try:
            call_command('import_books_xlsx', tmp_path)
            self.assertTrue(Book.objects.filter(isbn='9781234567897').exists())
        finally:
            import os

            os.remove(tmp_path)

    def test_import_parses_string_row_index_and_bare_isbn(self):
        tmp_path = _create_tmp_xlsx(
            [
                (
                    '2',
                    '978-0-12345-678-9, Algorithms and Data Structures, Some Author',
                ),
            ]
        )

        try:
            call_command('import_books_xlsx', tmp_path)
            self.assertTrue(Book.objects.filter(isbn='9780123456789').exists())
        finally:
            import os

            os.remove(tmp_path)
