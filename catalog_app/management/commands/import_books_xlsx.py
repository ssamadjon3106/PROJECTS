import re
from pathlib import Path
from typing import Any, Optional

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from library.models import Book


# Matches either:
# - "ISBN № 978-...-4"
# - or a bare ISBN value "978-...-4"
ISBN_ANY_RE = re.compile(r'(?i)(?:ISBN\s*№?\s*)?([0-9Xx-]{10,20})')
PREFIX_RE = re.compile(r'^(Книги[- ]?(?:Учебники|учебники)?.*?)(?:,|$)\s*')
PUBLISHERS = ('Pearson', 'Peason', 'Cengage', 'Cambridge', 'Oxford', 'Pan Macmillan')
SKIP_KEYWORDS = (
    'оборудования',
    'Шкаф',
    'Аппаратно-программный комплекс',
)


def compact(value: Any) -> str:
    return ' '.join(str(value or '').replace('\n', ' ').split())


def normalize_row_number(value: Any) -> Optional[int]:
    """
    Excel often stores the "index" column as float (e.g. 1.0) or as a string.
    Return a normalized integer row number or None if it can't be parsed.
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None

        # "3.0" -> 3
        if '.' in stripped:
            try:
                as_float = float(stripped)
            except ValueError:
                return None
            if as_float.is_integer():
                return int(as_float)
            return None

        try:
            return int(stripped)
        except ValueError:
            return None

    return None


def clean_isbn(value: Any, row_number: int) -> str:
    match = ISBN_ANY_RE.search(compact(value))
    if match:
        # Remove hyphens and spaces, normalize to uppercase.
        normalized = match.group(1).replace('-', '').strip().upper()
        return normalized

    # Fallback: keep uniqueness even if ISBN can't be extracted.
    return f'XLSX-{row_number:04d}'


def remove_prefix(value: str) -> str:
    return PREFIX_RE.sub('', value).strip(' ,')


def infer_category(value: str) -> str:
    lower = value.lower()
    if 'художественная' in lower or 'научно популярная' in lower:
        return 'Popular Science'
    if 'chemistry' in lower or 'biochemistry' in lower or 'chemical' in lower:
        return 'Chemistry'
    if 'physics' in lower:
        return 'Physics'
    if 'engineering' in lower or 'circuits' in lower or 'semiconductors' in lower:
        return 'Engineering'
    if 'economics' in lower or 'accounting' in lower or 'business' in lower or 'management' in lower:
        return 'Business'
    if 'computer' in lower or 'software' in lower or 'java' in lower or 'database' in lower:
        return 'Computer Science'
    if 'calculus' in lower or 'algebra' in lower or 'mathematics' in lower:
        return 'Mathematics'
    return 'English Textbooks'


def parse_book(row_number: int, description: Any) -> Optional[dict]:
    raw = compact(description)
    if not raw or any(keyword.lower() in raw.lower() for keyword in SKIP_KEYWORDS):
        return None

    isbn = clean_isbn(raw, row_number)

    text = remove_prefix(raw)
    # Remove the ISBN token whether it was labeled or bare.
    text = ISBN_ANY_RE.sub('', text).strip(' ,')

    publisher = ''
    for candidate in PUBLISHERS:
        if candidate.lower() in text.lower():
            publisher = 'Pearson' if candidate == 'Peason' else candidate
            break

    for candidate in PUBLISHERS:
        text = re.sub(rf'\b{re.escape(candidate)}\b', '', text, flags=re.IGNORECASE).strip(' ,')

    parts = [part.strip(' .') for part in text.split(',') if part.strip(' .')]
    if not parts:
        return None

    title_parts = [parts[0]]
    if len(parts) > 1 and ('edition' in parts[1].lower() or parts[1].lower().endswith('e')):
        title_parts.append(parts[1])

    title = ', '.join(title_parts)[:180]
    author_parts = parts[len(title_parts):]
    author = ', '.join(author_parts)[:120] or publisher or 'Unknown'

    return {
        'title': title,
        'author': author,
        'isbn': isbn[:20],
        'category': infer_category(raw),
        'shelf_location': f'Imported #{row_number}',
        'total_copies': 1,
        'available_copies': 1,
    }


class Command(BaseCommand):
    help = 'Import books from an Excel file with an index column and a description column.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Path to .xlsx file')
        parser.add_argument('--sheet', default=None, help='Optional sheet name')

    def handle(self, *args, **options):
        path = Path(options['path']).expanduser()
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        workbook = load_workbook(path, read_only=True, data_only=True)
        if options['sheet']:
            if options['sheet'] not in workbook.sheetnames:
                raise CommandError(f'Sheet "{options["sheet"]}" was not found.')
            worksheet = workbook[options['sheet']]
        else:
            worksheet = workbook.active

        created = 0
        updated = 0
        skipped = 0

        for row in worksheet.iter_rows(values_only=True):
            if not row or len(row) < 2:
                skipped += 1
                continue

            row_number_raw = row[0]
            description = row[1]

            row_number = normalize_row_number(row_number_raw)
            if row_number is None:
                skipped += 1
                continue

            data = parse_book(row_number, description)
            if not data:
                skipped += 1
                continue

            _, was_created = Book.objects.update_or_create(
                isbn=data['isbn'],
                defaults=data,
            )
            created += int(was_created)
            updated += int(not was_created)

        self.stdout.write(
            self.style.SUCCESS(
                f'Import complete from {path.name}: {created} created, {updated} updated, {skipped} skipped.'
            )
        )
